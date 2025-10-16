import tkinter as tk
from tkinter import messagebox
import traceback
import os
import sys

def run_fill_pereryvy():
    try:
        import pandas as pd
        from openpyxl import load_workbook

        # Определяем папку, где лежит EXE или .py
        if getattr(sys, 'frozen', False):
            folder = os.path.dirname(sys.executable)
        else:
            folder = os.path.dirname(os.path.abspath(__file__))

        # Отладка: выводим список файлов в папке
        file_list = os.listdir(folder)
        messagebox.showinfo("Отладка", f"Файлы в папке:\n{chr(10).join(file_list)}")

        pereryvy_path = os.path.join(folder, 'Перерывы.xlsx')
        tech_path = os.path.join(folder, '1ОБЩ_ТАБЛ_обновленный_тех_данные_по_МКД_1.xlsx')

        table_path = None
        checked_files = []
        for fname in file_list:
            checked_files.append(fname)
            if fname.lower().startswith('таблица_перерывов g') and fname.lower().endswith('.xlsx'):
                table_path = os.path.join(folder, fname)
                messagebox.showinfo("Файл найден", f"Файл для импорта: {fname}")
                break
        if table_path is None:
            messagebox.showerror("Ошибка поиска", f"Проверенные файлы:\n{chr(10).join(checked_files)}")
            raise FileNotFoundError('Не найден файл, начинающийся с "Таблица_перерывов g"')

        df = pd.read_excel(table_path)
        df_tech = pd.read_excel(tech_path, header=1)
        df_tech = df_tech.rename(columns={
            'Unnamed: 0': 'Адрес',
            'Unnamed: 2': 'Ду',
            'Unnamed: 3': 'ФИАС'
        })
        df_tech['Адрес'] = df_tech['Адрес'].astype(str).str.strip()
        tech_dict = df_tech.set_index('Адрес')[['Ду', 'ФИАС']].to_dict(orient='index')

        wb = load_workbook(pereryvy_path)
        sheet1 = wb['Информация о перерывах']
        sheet2 = wb['ОЖФ в инф. о перерывах']

        for ws in [sheet1, sheet2]:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

        start_row = 2
        for i, row in df.iterrows():
            sheet1.cell(row=start_row + i, column=1, value=row['Адрес'])
            sheet1.cell(row=start_row + i, column=2, value=row['Тип основания'])
            sheet1.cell(row=start_row + i, column=4, value=row['Вид коммунальной услуги'])
            sheet1.cell(row=start_row + i, column=5, value=row['Вид тарифицируемого ресурса'])
            sheet1.cell(row=start_row + i, column=6, value=row['Дата и время начала перерыва'])
            sheet1.cell(row=start_row + i, column=7, value=row['Дата и время окончания перерыва'])
            sheet1.cell(row=start_row + i, column=8, value=row['Причина перерыва'])
            sheet1.cell(row=start_row + i, column=9, value="Информация размещена")

        for i, row in df.iterrows():
            sheet2.cell(row=start_row + i, column=1, value=row['Адрес'])

        for row in range(2, sheet1.max_row + 1):
            adres_pereryv = sheet1.cell(row=row, column=1).value
            if adres_pereryv:
                adres_pereryv = str(adres_pereryv).strip()
                if adres_pereryv in tech_dict:
                    du_value = tech_dict[adres_pereryv]['Ду']
                    sheet1.cell(row=row, column=3, value=du_value)

        for row in range(2, sheet2.max_row + 1):
            adres_ojf = sheet2.cell(row=row, column=1).value
            if adres_ojf:
                adres_ojf = str(adres_ojf).strip()
                if adres_ojf in tech_dict:
                    fias_value = tech_dict[adres_ojf]['ФИАС']
                    sheet2.cell(row=row, column=3, value=fias_value)

        wb.save(pereryvy_path)
        messagebox.showinfo("Успех", "Файл успешно заполнен!")

    except Exception as e:
        err = traceback.format_exc()
        messagebox.showerror("Ошибка", f"Что-то пошло не так:\n{e}\n\nПодробности:\n{err}")

root = tk.Tk()
root.title("Автоматическое заполнение 'Перерывы'")

root.geometry("350x180")
label = tk.Label(root, text="Для заполнения файла 'Перерывы.xlsx'\nположите все нужные файлы в одну папку\nи нажмите кнопку ниже.", pady=20)
label.pack()

btn = tk.Button(root, text="Заполнить Перерывы", command=run_fill_pereryvy, height=2, width=30)
btn.pack()

root.mainloop()